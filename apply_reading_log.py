"""
apply_reading_log.py
--------------------
Reads a decisions JSON exported by my-book-selector.html and applies it to
both spreadsheets:

  1. books_output.xlsx    -> sets column Y ("Read") to "Yes" for any decision
                            with status Read / Ignored / Removed (i.e. anything
                            that should stop appearing in the selector).
  2. my-reading-log.xlsx  -> appends a row (Title, Author, Genre, Status,
                            Date in DD/MM/YYYY) for each decision, skipping
                            duplicates that already exist with the same status.

Match priority for finding a book in books_output.xlsx:
    ASIN  ->  ISBN_13  ->  ISBN_10  ->  DuplicateKey  ->  (title + author)

JSON format expected (produced by the modified selector):
{
  "exported_at": "2026-04-28T15:30:00",
  "decisions": [
    {
      "asin": "B0CTCQ6MM4",
      "isbn_13": "9781920000000",
      "isbn_10": "1917190077",
      "duplicate_key": "unbind|",
      "title": "UNBIND",
      "author": "Adam Wright",
      "genre": "Romance, Contemporary",
      "decision": "Read"          // Read | Ignored | Removed
    }
  ]
}

Usage:
    python apply_reading_log.py [path-to-decisions.json]

If no path is given, the most recent file in ~/Downloads matching
"book-selector-decisions*.json" is used.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding issues
if sys.platform.startswith("win"):
    try:
        import codecs
        sys.stdout.reconfigure(encoding='utf-8')
    except (AttributeError, OSError):
        # Python < 3.7 or reconfigure failed; wrap stdout
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

import openpyxl

# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(r"C:\Users\megha\OneDrive\Documents\Reading")


def _data_file(name: str) -> Path:
    """Spreadsheets live in the data folder; fall back to beside the code."""
    candidate = DATA_DIR / name
    return candidate if candidate.exists() else SCRIPT_DIR / name


BOOKS_OUTPUT = _data_file("books_output.xlsx")
READING_LOG = _data_file("my-reading-log.xlsx")

# 1-based column index in books_output.xlsx for the "Read" column.
# Column Y == column 25.  We verify the header on load.
READ_COL_INDEX = 25
READ_COL_HEADER = "Read"

# Records what you chose in the selector, kept apart from "Read" so that
# ignoring or removing a book no longer claims you read it.  Every non-empty
# value means "stop offering me this book"; only "Read" also sets Read = Yes.
DECISION_COL_HEADER = "Selector Decision"

VALID_DECISIONS = {"Read", "Ignored", "Removed"}
# Decisions that mean you actually finished the book.
READ_DECISIONS = {"Read"}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def find_decisions_file(arg: str | None) -> Path:
    if arg:
        p = Path(arg).expanduser().resolve()
        if not p.exists():
            sys.exit(f"ERROR: file not found: {p}")
        return p

    downloads = Path.home() / "Downloads"
    if not downloads.exists():
        sys.exit("ERROR: no path given and ~/Downloads doesn't exist.")
    candidates = sorted(
        downloads.glob("book-selector-decisions*.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        sys.exit(
            "ERROR: no decisions JSON given and no "
            "book-selector-decisions*.json found in Downloads."
        )
    return candidates[0]


def load_decisions(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    decisions = data.get("decisions") or []
    if not isinstance(decisions, list):
        sys.exit("ERROR: decisions JSON has no 'decisions' array.")
    cleaned: list[dict] = []
    for d in decisions:
        decision = (d.get("decision") or "").strip().capitalize()
        if decision not in VALID_DECISIONS:
            print(f"  skip (unknown decision={decision!r}): {d.get('title')}")
            continue
        cleaned.append({**d, "decision": decision})
    return cleaned


def norm(value) -> str:
    return ("" if value is None else str(value)).strip().lower()


def update_books_output(decisions: list[dict]) -> tuple[int, int]:
    """Mark each decided book as Read=Yes in books_output.xlsx column Y.
    Returns (updated_count, missing_count)."""

    if not BOOKS_OUTPUT.exists():
        print(f"  WARN: {BOOKS_OUTPUT.name} not found — skipping catalog update.")
        return 0, len(decisions)

    wb = openpyxl.load_workbook(BOOKS_OUTPUT)
    ws = wb.active

    # Verify column Y header to catch schema drift
    header_val = ws.cell(row=1, column=READ_COL_INDEX).value
    if header_val != READ_COL_HEADER:
        print(
            f"  WARN: column Y header is {header_val!r}, expected {READ_COL_HEADER!r}. "
            "Update will still proceed, but double-check your column order."
        )

    # Map each header to its 1-based column index
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col_idx).value
        if h:
            headers[str(h).strip()] = col_idx

    def col(name: str) -> int | None:
        return headers.get(name)

    # Build lookup indexes from the catalog
    asin_idx: dict[str, int] = {}
    isbn13_idx: dict[str, int] = {}
    isbn10_idx: dict[str, int] = {}
    dupkey_idx: dict[str, int] = {}
    title_author_idx: dict[tuple[str, str], int] = {}

    c_asin = col("ASIN")
    c_isbn13 = col("ISBN_13")
    c_isbn10 = col("ISBN_10")
    c_dupkey = col("DuplicateKey")
    c_title = col("Title")
    c_author = col("Author")

    for r in range(2, ws.max_row + 1):
        if c_asin and ws.cell(row=r, column=c_asin).value:
            asin_idx[norm(ws.cell(row=r, column=c_asin).value)] = r
        if c_isbn13 and ws.cell(row=r, column=c_isbn13).value:
            isbn13_idx[norm(ws.cell(row=r, column=c_isbn13).value)] = r
        if c_isbn10 and ws.cell(row=r, column=c_isbn10).value:
            isbn10_idx[norm(ws.cell(row=r, column=c_isbn10).value)] = r
        if c_dupkey and ws.cell(row=r, column=c_dupkey).value:
            dupkey_idx[norm(ws.cell(row=r, column=c_dupkey).value)] = r
        if c_title and c_author:
            t = norm(ws.cell(row=r, column=c_title).value)
            a = norm(ws.cell(row=r, column=c_author).value)
            if t or a:
                title_author_idx[(t, a)] = r

    # The decision column is new, so add it the first time this runs.
    c_decision = col(DECISION_COL_HEADER)
    if c_decision is None:
        c_decision = ws.max_column + 1
        ws.cell(row=1, column=c_decision).value = DECISION_COL_HEADER
        print(f"  Added '{DECISION_COL_HEADER}' column at position {c_decision}")

    # Match each decision and update
    updated = 0
    missing = 0
    marked_read = 0

    for d in decisions:
        asin = norm(d.get("asin"))
        isbn13 = norm(d.get("isbn_13"))
        isbn10 = norm(d.get("isbn_10"))
        dupkey = norm(d.get("duplicate_key"))
        title = norm(d.get("title"))
        author = norm(d.get("author"))

        row_num = None
        if asin and asin in asin_idx:
            row_num = asin_idx[asin]
        elif isbn13 and isbn13 in isbn13_idx:
            row_num = isbn13_idx[isbn13]
        elif isbn10 and isbn10 in isbn10_idx:
            row_num = isbn10_idx[isbn10]
        elif dupkey and dupkey in dupkey_idx:
            row_num = dupkey_idx[dupkey]
        elif (title, author) in title_author_idx:
            row_num = title_author_idx[(title, author)]

        if row_num:
            ws.cell(row=row_num, column=c_decision).value = d["decision"]
            # Only a genuine Read marks the book as read.  Ignored and Removed
            # stop it being offered again via the decision column instead.
            if d["decision"] in READ_DECISIONS:
                ws.cell(row=row_num, column=READ_COL_INDEX).value = "Yes"
                marked_read += 1
            updated += 1
        else:
            missing += 1
            print(f"  no match: {d.get('title')} by {d.get('author')}")

    wb.save(BOOKS_OUTPUT)
    print(f"  [OK] Recorded {updated} decisions in {BOOKS_OUTPUT.name} "
          f"({marked_read} also marked Read=Yes)")
    if missing:
        print(f"  [WARN] Could not find {missing} books")

    return updated, missing


def update_reading_log(decisions: list[dict]) -> int:
    """Append new rows to my-reading-log.xlsx for each decision.
    Skips duplicates (same title+author+status already exists).
    Returns count of rows appended."""

    if not READING_LOG.exists():
        print(f"  WARN: {READING_LOG.name} not found — skipping log update.")
        return 0

    wb = openpyxl.load_workbook(READING_LOG)
    ws = wb.active

    # Expected columns: Title, Author, Genre, Status, Date
    # Build a set of existing (title, author, status) tuples to skip duplicates
    existing: set[tuple[str, str, str]] = set()
    for r in range(2, ws.max_row + 1):
        t = norm(ws.cell(row=r, column=1).value)
        a = norm(ws.cell(row=r, column=2).value)
        s = norm(ws.cell(row=r, column=4).value)
        if t or a or s:
            existing.add((t, a, s))

    appended = 0
    today = datetime.now().strftime("%d/%m/%Y")

    for d in decisions:
        title = (d.get("title") or "").strip()
        author = (d.get("author") or "").strip()
        genre = (d.get("genre") or "").strip()
        status = d["decision"]  # Already capitalized: Read, Ignored, Removed

        key = (norm(title), norm(author), norm(status))
        if key in existing:
            continue  # Skip duplicate

        ws.append([title, author, genre, status, today])
        existing.add(key)
        appended += 1

    wb.save(READING_LOG)
    print(f"  [OK] Appended {appended} rows to {READING_LOG.name}")
    return appended


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def backfill_decisions(apply_changes: bool) -> None:
    """Recover past decisions from my-reading-log.xlsx.

    Earlier versions set Read = Yes for Ignored and Removed too, so the catalog
    claims you read books you only dismissed.  The reading log kept the real
    status, so it can say which is which.  Rows whose Read = Yes came from
    somewhere else (finishing a book, or StoryGraph reporting it as read) have
    no log entry and are left alone.
    """
    if not BOOKS_OUTPUT.exists() or not READING_LOG.exists():
        print("  Need both books_output.xlsx and my-reading-log.xlsx — skipping.")
        return

    log_wb = openpyxl.load_workbook(READING_LOG, read_only=True)
    log_ws = log_wb.active
    logged: dict[tuple[str, str], str] = {}
    for r in log_ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        title, author, status = norm(r[0]), norm(r[1]), (r[3] or "")
        status = str(status).strip().capitalize()
        if status not in VALID_DECISIONS:
            continue
        # A later "Read" outranks an earlier dismissal of the same book.
        key = (title, author)
        if logged.get(key) not in READ_DECISIONS:
            logged[key] = status
    log_wb.close()
    print(f"  Read {len(logged)} decision(s) from {READING_LOG.name}")

    wb = openpyxl.load_workbook(BOOKS_OUTPUT)
    ws = wb.active

    headers = {str(ws.cell(row=1, column=c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    c_decision = headers.get(DECISION_COL_HEADER)
    if c_decision is None:
        c_decision = ws.max_column + 1
        ws.cell(row=1, column=c_decision).value = DECISION_COL_HEADER
        print(f"  Added '{DECISION_COL_HEADER}' column at position {c_decision}")
    c_title, c_author = headers.get("Title"), headers.get("Author")
    if not c_title or not c_author:
        print("  Could not find Title/Author columns — skipping.")
        return

    tagged = cleared = 0
    for r in range(2, ws.max_row + 1):
        key = (norm(ws.cell(row=r, column=c_title).value),
               norm(ws.cell(row=r, column=c_author).value))
        status = logged.get(key)
        if not status:
            continue
        if not norm(ws.cell(row=r, column=c_decision).value):
            ws.cell(row=r, column=c_decision).value = status
            tagged += 1
        if status not in READ_DECISIONS:
            read_cell = ws.cell(row=r, column=READ_COL_INDEX)
            if norm(read_cell.value) == "yes":
                read_cell.value = ""
                cleared += 1

    print(f"  {tagged} row(s) tagged with a decision")
    print(f"  {cleared} row(s) had Read=Yes cleared (Ignored/Removed, not actually read)")

    if apply_changes:
        wb.save(BOOKS_OUTPUT)
        print(f"  [OK] Saved {BOOKS_OUTPUT.name}")
    else:
        print("  DRY RUN — nothing written. Re-run with --backfill --apply to save.")


def main() -> None:
    if "--backfill" in sys.argv:
        print(f"Backfilling decisions in {BOOKS_OUTPUT.name} from {READING_LOG.name}")
        backfill_decisions(apply_changes="--apply" in sys.argv)
        return

    path = find_decisions_file(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Reading decisions from: {path.name}")

    decisions = load_decisions(path)
    if not decisions:
        print("No valid decisions found in JSON.")
        sys.exit(0)

    print(f"Found {len(decisions)} decisions:")
    for d in decisions:
        print(f"  • {d['decision']}: {d.get('title')} by {d.get('author')}")

    print()
    updated, missing = update_books_output(decisions)
    appended = update_reading_log(decisions)

    print()
    print("=" * 60)
    print(f"Summary: {updated} catalog updates, {appended} log entries added")
    if missing:
        print(f"Warning: {missing} books not found in catalog")
    print("=" * 60)


if __name__ == "__main__":
    main()
